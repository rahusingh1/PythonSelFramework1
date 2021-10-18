import openpyxl

# locate the book from the particular location
book = openpyxl.load_workbook("E:\\PycharmProjects\\PythonSelFramework1\\TestData\\pythonDemo.xlsx")
sheet = book.active # point the active sheet of the book
cell = sheet.cell(row=1, column=1) # point the cell of the sheet
print(cell.value) # get and print the value from the cell

sheet.cell(row=2, column=2).value = "Rahul" # assing the value to particular cell
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row) # it prints the maximum no of rows in the sheet
print(sheet.max_column) # it prints the maximum no of columns in the sheet

print(sheet['A5'].value) # shortcut to print the value of that particular cell.

# for i in range(1, sheet.max_row+1): # to get rows
#     for j in range(1, sheet.max_column + 1): # to get columns
#         print(sheet.cell(row=i, column=j).value)

# if you have a requirement to print the values of testcase2 only
# for i in range(1, sheet.max_row+1): # to get rows
#     if sheet.cell(row=i, column=1).value == "Testcase2":
#         for j in range(1, sheet.max_column + 1): # to get columns
#             print(sheet.cell(row=i, column=j).value)

# to extract the data from here and store it in the dictionary to be used by the test cases.
Dict = {}  # initialize empty dictionary
for i in range(1, sheet.max_row+1): # to get rows
      if sheet.cell(row=i, column=1).value == "Testcase2":
          for j in range(2, sheet.max_column + 1): # to get columns
              Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)